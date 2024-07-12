import openpyxl
from tkinter import messagebox

class Profiles:
    def __init__(self, cy, cx, cm):
        self.cy = cy
        self.cx = cx
        self.cm = cm
        self.idial = []  # список для строк, полностью совпадающей с введёнными данными
        self.best = []  # список для строк, в которых 3 параметра лучше введённых
        self.good = []  # список для строк, в которых любые 2 параметра лучше введённых
        self.norm = []  # список для строк, в которых только 1 параметр лучше введённого
        self.bad = []  # список для строк, в которых все параметры хуже введённых
        self.a = int  # переменная, в которую будет перезаписываться наиближайшая строчка
        self.srav = 9999.9  # разность-ориентир, с которой сравниваем. очень большая, чтобы потом при сравнении с любой меньшей разностью последнюю сделать ориентиром
        self.dy = float
        self.dx = float
        self.dm = float
        self.d = float
        self.d1 = 0.0  # ориентир для сравнения в списке k1. Сравниваем с малым в поисках бóльшего значения
        self.d4 = 9999.9  # ориентир для сравнения в списке k3. Сравниваем с большим в поисках меньшего
        self.best_best = int
        self.bad_bad = int


    def relating(self, cy, cx, cm):
        wb = openpyxl.load_workbook('data2.xlsx')  # доступ к эксель странице
        ws = wb.active
        def calc(rate, row_ind, column_ind):
            return (float(ws.cell(row=row_ind, column=column_ind).value) - float(rate))/float(rate)


        for i in range(2, ws.max_row):  # пробегаем строчки
            if not ws.cell(row=i, column=5).value:
                continue
            self.dy = calc(rate=cy, row_ind=i,column_ind=3)*100  # относительное изменение Cy из таблицы относительно введённого в процентах. Положительное значение - коэффициент лучше введённого (больше)
            self.dx = calc(rate=cx, row_ind=i, column_ind=4)*(-100)  # относительное изменение Cx из таблицы относительно введённого в процентах. Положительное значение - коэффициент лучше введённого (меньше)
            self.dm = calc(rate=cm, row_ind=i, column_ind=5)*(-100)  # относительное изменение Cm из таблицы относительно введённого в процентах. Положительное значение - коэффициент лучше введённого (меньше)
            self.d = (self.dy**2 + self.dx**2 + self.dm**2)**0.5  # геометрическое расстояние между проверяемыми и введёнными данными

            #self.dy = round(self.dy, 3)  # для наглядности при отработке (потом можно будет вырезать вместе с процентами, когда закончится отладки кода)
            #self.dx = round(self.dx, 3)
            #self.dm = round(self.dm, 3)
            #print(i, "dy = ", self.dy, "dx = ", self.dx, "dm = ", self.dm, "d = ", self.d)  # тоже для проверки

            if self.dy > 0 and self.dx > 0 and self.dm > 0:  # список best, если строка подходит по значениям, номер строки записывается в список
                self.best.append(i)
                if self.d > self.d1:  # для 1-го списка можно находить наибольшее расстояние между точками, для 3-го - наименьшее
                    self.d1 = self.d  # чтобы среди списка найти наилучшие показатели
                    self.best_best = i  # номер строки с более хорошими коэффициентами в 1-м списке
            if (self.dy > 0 and self.dx > 0 and self.dm < 0) or (self.dy > 0 and self.dx < 0 and self.dm > 0) or (self.dy < 0 and self.dx > 0 and self.dm > 0):  # до 39 строки кода аналогично
                self.good.append(i)
            if (self.dy > 0 and self.dx < 0 and self.dm < 0) or (self.dy < 0 and self.dx < 0 and self.dm > 0) or (self.dy < 0 and self.dx > 0 and self.dm < 0):
                self.norm.append(i)
            if self.dy < 0 and self.dx < 0 and self.dm < 0:
                self.bad.append(i)
                if self.d < self.d4:
                    self.d4 = self.d
                    self.bad_bad = i
            if self.dy == 0 and self.dx == 0 and self.dm == 0:  # если находится полное совпадение, номер строки в отдельный список
                self.idial.append(i)

        #print(self.dy, self.dx, self.dm)
        # если 4-й список не пустой, присваиваем переменной а значение нужной строки
        if self.idial:
            self.a = self.idial[len(self.idial)-1]
        # если полного совпадения нет, просматриваем аналогично список с всеми коэффициентами выше введённых
        else:
            if self.best:
                self.a = self.best_best  # строка с наивысшими коэффициентами по отношению к введённым
        # если 1-й список пуст, проверяем с двумя коэффициентами выше введённых
            else:
                if self.good:
                    self.a = self.good[len(self.good) - 1]  # пока используется последняя записанная в список строка. Нужно продумать алгоритм!
        # аналогия с первым списком
                else:
                    if self.norm:
                        self.a = self.norm[len(self.norm) - 1]
                    else:
                        if self.bad:
                            self.a = self.bad_bad
                        else:
                            messagebox.showerror('Ошибка', 'Нет подходящих профилей')

        # Вывод данных
        print(ws.cell(row=self.a, column=1).value)  # профиль
        print(ws.cell(row=self.a, column=2).value)  # угол атаки
        print(ws.cell(row=self.a, column=3).value)  # Cy
        print(ws.cell(row=self.a, column=4).value)  # Cx
        print(ws.cell(row=self.a, column=5).value)  # Cm

example = Profiles(cy=float, cx=float, cm=float)
#example.relating(cy=float(input()), cx=float(input()), cm=float(input()))
print(example.relating(cy=float(input()), cx=float(input()), cm=float(input())))