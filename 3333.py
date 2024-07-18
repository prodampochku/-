import openpyxl

# вводим с клавиатуры коэффициенты
b_0 = input()  #корневая хорда
b_k = input()  #концевая хорда
l = input()  #размах крыла
v = input()  #скорость набегающего потока
Y = input()  #подъёмная сила
X = input()  #сила лобового сопротивления
Mz = input()  #продольный момент (тангажа)
rho = 0.125 #плотность воздуха при н.у.; константа
def str_to_num(line):  #функция переводит данное ей значение в строку (вместо line даём какую-либо переменную)
    line = line.strip()
    if line.isdigit():  #если значение состоит только из цифр, возвращает целочисленное значение
        return int(line)
    elif '.' in line:  #если встречается точка, возвращает вещественное значение
        return float(line)
    elif ',' in line:  #если встречается запятая, она заменяется на точку, с которой возвращается вещественное значение
        return float(line.replace(',', '.'))

b_0 = str_to_num(line=b_0)
b_k = str_to_num(line=b_k)
l = str_to_num(line=l)
v = str_to_num(line=v)
Y = str_to_num(line=Y)
X = str_to_num(line=X)
Mz = str_to_num(line=Mz)

b_mid = (b_0 + b_k)/2  #средняя хорда
s = 0.5 * (b_0 + b_k) * l  #площадь крыла
cy = 2*Y/(rho*s*(v**2))  #коэффициент подъёмной силы
cx = 2*X/(rho*s*(v**2))  #коэффициент лобового сопротивления
cm = 2*Mz/(rho*s*b_mid*(v**2))  #коэффициент момента тангажа

delta_cy = 9999.9  #ориентиры-сравнения
delta_cx = 9999.9
delta_cm = 9999.9
sp_cy = []  #списки для строк, подходящих по одному параметру
sp_cx = []
sp_cm = []

val = int  #переменная, которая хранит локальный номер строки
m = 9999.9 #переменная для сравнения в списках
m1 = float() #переменная для записи в списках
k1 = 1  #переменная, которая хранит номер строки с наиближайшими коэффициентами

wb = openpyxl.load_workbook('data.xlsx')
ws = wb.active

def calc(row_ind, column_ind, rate):  #разность введённого коэффициента и проверяемого. в функции используются параметры номера строки, столбца и коэффициента
    return (rate - float(ws.cell(row=row_ind, column=column_ind).value))


#ищем строк, которые теоритчески могут подойти, сохраняем номер строк из бд, которые всё ближе и ближе к вводимым данным
for i in range(2, ws.max_row):
        if not ws.cell(row=i, column=5).value:  #если ячейка cm в данной строке пустая, строка пропускается
            continue
        if calc(row_ind=i, column_ind=3, rate=cy) **2 - (delta_cy)**2 <= 0:  # разность целевого Cy и
                    # проверяемого возводим в квадрат и сравниваем с квадратом разности-ориентира
            delta_cy = abs(calc(i, 3, cy))  # если меньше разности-ориентира,
                    # то модуль разности целевого и проверяемого Cy становится новой разностью-ориентиром
            sp_cy.append(i)  # в список sp_cy добавляем номер строки, которая проверялась
        elif calc(row_ind=i, column_ind=3, rate=cy) **2 - (delta_cy)**2 > 0:  # если больше разности-ориентира, то ничего не делает программа
            pass
        if calc(row_ind=i, column_ind=4, rate=cx) **2 - (delta_cx)**2 <= 0:
            delta_cx = abs(calc(i, 4, cx))
            sp_cx.append(i)
        elif calc(row_ind=i, column_ind=4, rate=cx) **2 - (delta_cx)**2 > 0:
            pass
        if calc(row_ind=i, column_ind=5, rate=cm) **2 - (delta_cm)**2 <= 0:
            delta_cm = abs(calc(i, 5, cm))
            sp_cm.append(i)
        elif calc(row_ind=i, column_ind=5, rate=cm) **2 - (delta_cm)**2 > 0:
            pass

#ищем наиболее близкое геометрически (корень из суммы квадратов)
for i in range(0, len(sp_cy)): #прогон по списку с Cy
    val = sp_cy[i] #переменной val присваиваем i-е значение из списка sp_cy
    m1 = (((float(ws.cell(row=val, column=3).value) - float(cy))**2) + (
                    (float(ws.cell(row=val, column=4).value) - float(cx))** 2) + ((float(ws.cell(row=val, column=5).value) - float(cm))** 2))** 0.5
#переменной m1 присваиваем расстояние между точками. Координаты 1-й точки - коэффициенты проверяемой строки, 2-й - введённые коэффициенты
    if m1 < m: # если расстояние между точками меньше хрени для сравнения, то сравнивать будет
                # теперь с вычисленным расстоянием между точками (m - минимум, с которым сравниваем и ищем меньшее),
                # иначе ничего не делаем
        m = m1
        k1 = val #сохраняем лучший пока что номер строки

for j in range(0, len(sp_cx)): #аналогично проверяем списки с номерами строк близких cx и cm к введённым
    val = sp_cx[j]
    m1 = (((float(ws.cell(row=val, column=3).value) - float(cy)) ** 2) + (
                    (float(ws.cell(row=val, column=4).value) - float(cx)) ** 2) + ((float(ws.cell(row=val, column=5).value) - float(cm)) ** 2)) ** 0.5
    if m1 < m:
        m = m1
        k1 = val

for k in range(0, len(sp_cm)):
    val = sp_cm[k]
    m1 = (((float(ws.cell(row=val, column=3).value) - float(cy)) ** 2) + (
                    (float(ws.cell(row=val, column=4).value) - float(cx)) ** 2) + ((float(ws.cell(row=val, column=5).value) - float(cm)) ** 2)) ** 0.5
    if m1 < m:
        m = m1
        k1 = val

delta_cy = 9999.9  #возвращаем исходные значения для ориенторив
delta_cx = 9999.9
delta_cm = 9999.9
m = 9999.9
k1 = 1
sp_cy = sp_cy.clear()  #списки очищаем
sp_cx = sp_cx.clear()
sp_cm = sp_cm.clear()

#вывод
print("Наиближайший профиль")
print(ws.cell(row=k1, column=1).value)  # профиль
print(ws.cell(row=k1, column=2).value)  # угол атаки
print(ws.cell(row=k1, column=3).value)  # Cy
print(ws.cell(row=k1, column=4).value)  # Cx
print(ws.cell(row=k1, column=5).value)  # Cm

