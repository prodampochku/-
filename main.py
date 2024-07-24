from PyQt5 import QtWidgets, uic
from PyQt5.QtGui import QPixmap
import sys

import matplotlib
import openpyxl

app = QtWidgets.QApplication([])

ui = uic.loadUi('design.ui')

list_angle = []  # Заполняем углы атаки
for i in range(0, 100, 10):
    list_angle.append(str(i))

ui.alpha_Box.addItems(list_angle)

ui.label_9.setStyleSheet('color: rgb(0, 139, 210)')
ui.textEdit.setStyleSheet('color: rgb(0, 139, 210)')
ui.label_10.setStyleSheet('color: rgb(0, 139, 210)')
ui.label_2.setStyleSheet('color: rgb(0, 139, 210)')
ui.label_3.setStyleSheet('color: rgb(0, 139, 210)')
ui.label.setStyleSheet('color: rgb(0, 139, 210)')
ui.label_4.setStyleSheet('color: rgb(0, 139, 210)')
ui.label_5.setStyleSheet('color: rgb(0, 139, 210)')
ui.label_6.setStyleSheet('color: rgb(0, 139, 210)')
ui.label_7.setStyleSheet('color: rgb(0, 139, 210)')
ui.label_8.setStyleSheet('color: rgb(0, 139, 210)')
ui.label_11.setStyleSheet('color: rgb(0, 139, 210)')
ui.label_12.setStyleSheet('color: rgb(0, 139, 210)')
ui.label_15.setStyleSheet('color: rgb(0, 139, 210)')
ui.label_16.setStyleSheet('color: rgb(0, 139, 210)')
ui.label_17.setStyleSheet('color: rgb(0, 139, 210)')

ui.geometry.setColumnCount(3)
ui.geometry.setRowCount(25)


def algoritm():
    # вводим с клавиатуры коэффициенты
    b_0 = ui.finall_horda.text()  # корневая хорда
    b_k = ui.current_horda.text()  # концевая хорда
    l = ui.razmah.text()  # размах крыла
    v = ui.V_wing.text()  # скорость набегающего потока
    Y = ui.Cy_current.text()  # подъёмная сила
    X = ui.Cx_current.text()  # сила лобового сопротивления
    Mz = ui.Mz_current.text()  # продольный момент (тангажа)
    rho = 1.225  # плотность воздуха при н.у.; константа

    def str_to_num(line):  # функция переводит данное ей значение в строку (вместо line даём какую-либо переменную)
        line = line.strip()
        if line.isdigit():  # если значение состоит только из цифр, возвращает целочисленное значение
            return int(line)
        elif '.' in line:  # если встречается точка, возвращает вещественное значение
            return float(line)
        elif ',' in line:  # если встречается запятая, она заменяется на точку, с которой возвращается вещественное значение
            return float(line.replace(',', '.'))

    b_0 = str_to_num(line=b_0)
    b_k = str_to_num(line=b_k)
    l = str_to_num(line=l)
    v = str_to_num(line=v)
    Y = str_to_num(line=Y)
    X = str_to_num(line=X)
    Mz = str_to_num(line=Mz)

    b_mid = (b_0 + b_k) / 2  # средняя хорда
    s = 0.5 * (b_0 + b_k) * l  # площадь крыла
    cy = 2 * Y / (rho * s * (v ** 2))  # коэффициент подъёмной силы
    cx = 2 * X / (rho * s * (v ** 2))  # коэффициент лобового сопротивления
    cm = 2 * Mz / (rho * s * b_mid * (v ** 2))  # коэффициент момента тангажа

    delta_cy = 9999.9  # ориентиры-сравнения
    delta_cx = 9999.9
    delta_cm = 9999.9
    sp_cy = []  # списки для строк, подходящих по одному параметру
    sp_cx = []
    sp_cm = []

    val = int  # переменная, которая хранит локальный номер строки
    m = 9999.9  # переменная для сравнения в списках
    m1 = float()  # переменная для записи в списках
    k1 = 1  # переменная, которая хранит номер строки с наиближайшими коэффициентами
    k2 = 0  # переменная, которая хранит номер столбца нужного профиля на листе с геометрией

    wb = openpyxl.load_workbook('data.xlsx')  #страницы в эксель с именами: коэффициенты отечественных профилей, иностранных (пока пусто) и геометрия отечественных профилей
    aerodym_sheet = wb["адх"]
    aerodym_2_sheet = wb["адх2"]
    geometry_sheet = wb["геометрия"]

    def calc(row_ind, column_ind,
             rate):  # разность введённого коэффициента и проверяемого. в функции используются параметры номера строки, столбца и коэффициента
        return (rate - float(aerodym_sheet.cell(row=row_ind, column=column_ind).value))

    # ищем строк, которые теоритчески могут подойти, сохраняем номер строк из бд, которые всё ближе и ближе к вводимым данным
    for i in range(2, aerodym_sheet.max_row):
        if not aerodym_sheet.cell(row=i, column=5).value:  # если ячейка cm в данной строке пустая, строка пропускается
            continue
        if calc(row_ind=i, column_ind=3, rate=cy) ** 2 - (delta_cy) ** 2 <= 0:  # разность целевого Cy и
            # проверяемого возводим в квадрат и сравниваем с квадратом разности-ориентира
            delta_cy = abs(calc(i, 3, cy))  # если меньше разности-ориентира,
            # то модуль разности целевого и проверяемого Cy становится новой разностью-ориентиром
            sp_cy.append(i)  # в список sp_cy добавляем номер строки, которая проверялась
        elif calc(row_ind=i, column_ind=3, rate=cy) ** 2 - (
        delta_cy) ** 2 > 0:  # если больше разности-ориентира, то ничего не делает программа
            pass
        if calc(row_ind=i, column_ind=4, rate=cx) ** 2 - (delta_cx) ** 2 <= 0:
            delta_cx = abs(calc(i, 4, cx))
            sp_cx.append(i)
        elif calc(row_ind=i, column_ind=4, rate=cx) ** 2 - (delta_cx) ** 2 > 0:
            pass
        if calc(row_ind=i, column_ind=5, rate=cm) ** 2 - (delta_cm) ** 2 <= 0:
            delta_cm = abs(calc(i, 5, cm))
            sp_cm.append(i)
        elif calc(row_ind=i, column_ind=5, rate=cm) ** 2 - (delta_cm) ** 2 > 0:
            pass

    # ищем наиболее близкое геометрически (корень из суммы квадратов)
    for i in range(0, len(sp_cy)):  # прогон по списку с Cy
        val = sp_cy[i]  # переменной val присваиваем i-е значение из списка sp_cy
        m1 = (((float(aerodym_sheet.cell(row=val, column=3).value) - float(cy)) ** 2) + (
                    (float(aerodym_sheet.cell(row=val, column=4).value) - float(cx)) ** 2) + (
                          (float(aerodym_sheet.cell(row=val, column=5).value) - float(cm)) ** 2)) ** 0.5
        # переменной m1 присваиваем расстояние между точками. Координаты 1-й точки - коэффициенты проверяемой строки, 2-й - введённые коэффициенты
        if m1 < m:  # если расстояние между точками меньше хрени для сравнения, то сравнивать будет
            # теперь с вычисленным расстоянием между точками (m - минимум, с которым сравниваем и ищем меньшее),
            # иначе ничего не делаем
            m = m1
            k1 = val  # сохраняем лучший пока что номер строки

    for j in range(0, len(sp_cx)):  # аналогично проверяем списки с номерами строк близких cx и cm к введённым
        val = sp_cx[j]
        m1 = (((float(aerodym_sheet.cell(row=val, column=3).value) - float(cy)) ** 2) + (
                    (float(aerodym_sheet.cell(row=val, column=4).value) - float(cx)) ** 2) + (
                          (float(aerodym_sheet.cell(row=val, column=5).value) - float(cm)) ** 2)) ** 0.5
        if m1 < m:
            m = m1
            k1 = val

    for k in range(0, len(sp_cm)):
        val = sp_cm[k]
        m1 = (((float(aerodym_sheet.cell(row=val, column=3).value) - float(cy)) ** 2) + (
                    (float(aerodym_sheet.cell(row=val, column=4).value) - float(cx)) ** 2) + (
                          (float(aerodym_sheet.cell(row=val, column=5).value) - float(cm)) ** 2)) ** 0.5
        if m1 < m:
            m = m1
            k1 = val

    # вывод
    # print("Наиближайший профиль")
    ui.name_profile_print.setText(
        str((aerodym_sheet.cell(row=k1, column=1).value)))  # print(aerodym_sheet.cell(row=k1, column=1).value)  # профиль
    # print(ws.cell(row=k1, column=2).value)  # угол атаки
    ui.Cy_finall.setText(str((aerodym_sheet.cell(row=k1, column=3).value)))  # Cy
    ui.Cx_finall.setText(str((aerodym_sheet.cell(row=k1, column=4).value)))  # Cx
    ui.Mz_finall.setText(str((aerodym_sheet.cell(row=k1, column=5).value)))  # Cm
    pixmap = QPixmap('p5/Profiles/' + (aerodym_sheet.cell(row=k1, column=6).value))
    ui.label_14.setPixmap(pixmap)


    for i in range(1, geometry_sheet.max_column, 3):  #на 3 страницы проходим циклом через каждые 3 столбца
        if str(geometry_sheet.cell(row=1, column=i).value) == str(aerodym_sheet.cell(row=k1, column=1).value):
            k2 = i  #если значение ячейки (название профиля) совпадает с тем, что мы нашли и вывели, то запоминаем номер столбца и заканчиваем цикл
            break

    a=[]
    for i in range(3 , geometry_sheet.max_row):  #с 3 строчки, где начинаются координаты точек, до последней строчкив столбце цикл
        if geometry_sheet.cell(row=i, column=1).value == None:  #если ячейка в данной строке пустая, строка пропускается
            continue
        for j in range(k2, k2 + 3):  #выводим k2-й столбец и 2 следующих за ним
            print(geometry_sheet.cell(row=i, column=j).value, end='  ')  # end='  ' убрал перенос каждого значения на новую строку
            ui.geometry.setItem(i-3, j-k2, QTableWidgetItem(str(geometry_sheet.cell(row=i, column=j).value)))
        print()  #а этот принт завершал строчку. Т.к. вывод в приложении в таблицу, то здесь ты, Маша, лишние принты уберёшь сама


ui.btn1.clicked.connect(algoritm)
ui.Tab_widget.setCurrentIndex(0)


def tab_switch1():
    ui.Tab_widget.setCurrentIndex(1)


ui.start_btn.clicked.connect(tab_switch1)


def tab_switch2():
    ui.Tab_widget.setCurrentIndex(2)


ui.btn1.clicked.connect(tab_switch2)
ui.show()
app.exec()
