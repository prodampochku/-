import openpyxl
from tkinter import messagebox

class Profiles:
    def __init__(self, cy, cx, cm):
        self.cy = cy
        self.cx = cx
        self.cm = cm

    def relating(self, cy, cx, cm):
        self.idial = []  # список для строк, полностью совпадающей с введёнными данными
        self.best = []  # список для строк, в которых 3 параметра лучше введённых
        self.good = []  # список для строк, в которых любые 2 параметра лучше введённых
        self.norm = []  # список для строк, в которых только 1 параметр лучше введённого
        self.bad = []  # список для строк, в которых все параметры хуже введённых
        self.a = int  # переменная, в которую будет перезаписываться наиближайшая строчка
        self.srav = 9999.9  # разность-ориентир, с которой сравниваем. очень большая, чтобы потом при сравнении с любой меньшей разностью последнюю сделать ориентиром
        self.dy = float
        self.dx = float
        self.dm = float
        self.d = float
        self.d1 = 0.0  # ориентир для сравнения в списке best. Сравниваем с малым в поисках бóльшего значения
        self.d4 = 9999.9  # ориентир для сравнения в списке bada. Сравниваем с большим в поисках меньшего
        self.best_num = int
        self.bad_num = int

        wb = openpyxl.load_workbook('data2.xlsx')  # доступ к эксель странице
        ws = wb.active
        def calc(rate, row_ind, column_ind):
            return (float(ws.cell(row=row_ind, column=column_ind).value) - float(rate))/float(rate)

        for i in range(2, ws.max_row):  # пробегаем строчки
            if not ws.cell(row=i, column=5).value:
                continue
            self.dy = calc(rate=cy, row_ind=i,column_ind=3)  # относительное изменение Cy из таблицы относительно введённого в процентах. Положительное значение - коэффициент лучше введённого (больше)
            self.dx = calc(rate=cx, row_ind=i, column_ind=4)*(-1)  # относительное изменение Cx из таблицы относительно введённого в процентах. Положительное значение - коэффициент лучше введённого (меньше)
            self.dm = calc(rate=cm, row_ind=i, column_ind=5)*(-1)  # относительное изменение Cm из таблицы относительно введённого в процентах. Положительное значение - коэффициент лучше введённого (меньше)
            self.d = (self.dy**2 + self.dx**2 + self.dm**2)**0.5  # геометрическое расстояние между проверяемыми и введёнными данными

            #self.dy = round(self.dy, 3)  # для наглядности при отработке (потом можно будет вырезать вместе с процентами, когда закончится отладки кода)
            #self.dx = round(self.dx, 3)
            #self.dm = round(self.dm, 3)
            #print(i, "dy = ", self.dy, "dx = ", self.dx, "dm = ", self.dm, "d = ", self.d)  # тоже для проверки

            if self.dy > 0 and self.dx > 0 and self.dm > 0:  # список best, если строка подходит по значениям, номер строки записывается в список
                self.best.append(i)
                if self.d > self.d1:  # для 1-го списка можно находить наибольшее расстояние между точками, для 3-го - наименьшее
                    self.d1 = self.d  # чтобы среди списка найти наилучшие показатели
                    self.best_num = i  # номер строки с более хорошими коэффициентами в 1-м списке
            if (self.dy > 0 and self.dx > 0 and self.dm < 0) or (self.dy > 0 and self.dx < 0 and self.dm > 0) or (self.dy < 0 and self.dx > 0 and self.dm > 0):  # до 39 строки кода аналогично
                self.good.append(i)
            if (self.dy > 0 and self.dx < 0 and self.dm < 0) or (self.dy < 0 and self.dx < 0 and self.dm > 0) or (self.dy < 0 and self.dx > 0 and self.dm < 0):
                self.norm.append(i)
            if self.dy < 0 and self.dx < 0 and self.dm < 0:
                self.bad.append(i)
                if self.d < self.d4:
                    self.d4 = self.d
                    self.bad_num = i
            if self.dy == 0 and self.dx == 0 and self.dm == 0:  # если находится полное совпадение, номер строки в отдельный список
                self.idial.append(i)

        # если 4-й список не пустой, присваиваем переменной а значение нужной строки
        if self.idial:
            self.a = self.idial[len(self.idial)-1]
        # если полного совпадения нет, просматриваем аналогично список с всеми коэффициентами выше введённых
        else:
            if self.best:
                self.a = self.best_num  # строка с наивысшими коэффициентами по отношению к введённым
        # если 1-й список пуст, проверяем с двумя коэффициентами выше введённых
            else:
                if self.good:
                    self.a = self.good[len(self.good) - 1]  # пока используется последняя записанная в список строка. Нужно продумать алгоритм!
        # аналогия с первым списком
                else:
                    if self.norm:
                        self.a = self.norm[len(self.norm) - 1]
                    else:
                        if self.bad:
                            self.a = self.bad_num
                        else:
                            messagebox.showerror('Ошибка', 'Нет подходящих профилей')

        # Вывод данных
        print("Наилучший профиль")
        print(ws.cell(row=self.a, column=1).value)  # профиль
        print(ws.cell(row=self.a, column=2).value)  # угол атаки
        print(ws.cell(row=self.a, column=3).value)  # Cy
        print(ws.cell(row=self.a, column=4).value)  # Cx
        print(ws.cell(row=self.a, column=5).value)  # Cm

    def sravn(self, cy, cx, cm, val): #пробная функция для приближённого поиска
        self.delta_cy = 9999.9
        self.delta_cx = 9999.9
        self.delta_cm = 9999.9
        self.sp_cy = []
        self.sp_cx = []
        self.sp_cm = []


        def calc2(row_ind, column_ind, rate):
            return (rate - float(ws.cell(row=row_ind, column=column_ind).value))

        wb = openpyxl.load_workbook('data2.xlsx')
        ws = wb.active
        #ищем строк, которые теоритчески могут подойти, сохраняем номер строк из бд, которые всё ближе и ближе к вводимым данным
        for i in range(2, ws.max_row):
                if not ws.cell(row=i, column=5).value:
                    continue
                if calc2(row_ind=i, column_ind=3, rate=cy)**2 - (self.delta_cy)**2 < 0:
                    self.delta_cy = abs(calc2(i, 3, cy))
                    self.sp_cy.append(i)
                elif calc2(row_ind=i, column_ind=3, rate=cy)**2 - (self.delta_cy)**2 >= 0:
                    pass
                if calc2(row_ind=i, column_ind=4, rate=cx)**2 - (self.delta_cx)**2 < 0:
                    self.delta_cx = abs(calc2(i, 4, cx))
                    self.sp_cx.append(i)
                elif calc2(row_ind=i, column_ind=4, rate=cx)**2 - (self.delta_cx)**2 >= 0:
                    pass
                if calc2(row_ind=i, column_ind=5, rate=cm)**2 - (self.delta_cm)**2 < 0:
                    self.delta_cm = abs(calc2(i, 5, cm))
                    self.sp_cm.append(i)
                elif calc2(row_ind=i, column_ind=5, rate=cm)**2 - (self.delta_cm)**2 >= 0:
                    pass

                #print(self.delta_cy, '||', self.delta_cx, '||', self.delta_cm, '||', self.sp_cy, '||', self.sp_cx, '||', self.sp_cm)

                '''if (cx - float(ws.cell(row=i, column=4).value))**2 < (self.dx)**2: # разность целевого Cx и
                    # проверяемого возводим в квадрат и сравниваем с квадратом разности-ориентира
                    self.dx = abs(cx - float(ws.cell(row=i, column=4).value)) # если меньше разности-ориентира,
                    # то модуль разности целевого и проверяемого Cx становится новой разностью-ориентиром
                    self.sp_cx.append(i) # в список sp_cx добавляем номер строки, которая проверялась
                elif (cx - float(ws.cell(row=i, column=4).value)) ** 2 >= (self.dx) ** 2: # если больше разности-ориентира, то ничего не делает программа
                    pass

                if (cm - float(ws.cell(row=i, column=5).value)) ** 2 < (self.dm) ** 2: # разность целевого Cm и
                    # проверяемого возводим в квадрат и сравниваем с квадратом разности-ориентира
                    self.dm = abs(cm - float(ws.cell(row=i, column=5).value)) # если меньше разности-ориентира,
                    # то модуль разности целевого и проверяемого Cm становится новой разностью-ориентиром
                    self.sp_cm.append(i) # в список sp_cm добавляем номер строки, которая проверялась
                elif (cm - float(ws.cell(row=i, column=5).value)) ** 2 >= (self.dm) ** 2: # если больше разности-ориентира, то ничего не делает программа
                    pass'''
        #print(self.sp_cy, self.sp_cx, self.sp_cm) #выводит номера строк, в которых были подходящие профили в порядке улучшения (3 списка для Сy, Cx, Cm)

        #ищем наиболее близкое геометрически (корень из суммы квадратов)
        self.val = val #переменная, которая хранит локальный номер строки
        self.m = 9999.9 #переменная для сравнения
        self.m1 = float() #переменная для записи
        self.k1 = 1 #переменная, которая хранит номер строки с наиближайшими коэффициентами
        for i in range(0, len(self.sp_cy)): #прогон по списку с Cy
            val = self.sp_cy[i] #переменной val присваиваем i-е значение из списка sp_cy
            self.m1 = (((float(ws.cell(row=val, column=3).value) - float(cy))**2) + (
                    (float(ws.cell(row=val, column=4).value) - float(cx))** 2) + ((float(ws.cell(row=val, column=5).value) - float(cm))** 2))** 0.5
            #переменной m1 присваиваем расстояние между точками. Координаты 1-й точки - коэффициенты проверяемой строки, 2-й - введённые коэффициенты
            if self.m1 < self.m: # если расстояние между точками меньше хрени для сравнения, то сравнивать будет
                # теперь с вычисленным расстоянием между точками (m - минимум, с которым сравниваем и ищем меньшее),
                # иначе ничего не делаем
                self.m = self.m1
                self.k1 = val #сохраняем лучший пока что номер строки

        for j in range(0, len(self.sp_cx)): #аналогично проверяем списки с номерами строк близких cx и cm к введённым
            val = self.sp_cx[j]
            self.m1 = (((float(ws.cell(row=val, column=3).value) - float(cy)) ** 2) + (
                    (float(ws.cell(row=val, column=4).value) - float(cx)) ** 2) + ((float(ws.cell(row=val, column=5).value) - float(cm)) ** 2)) ** 0.5
            if self.m1 < self.m:
                self.m = self.m1
                self.k1 = val

        for k in range(0, len(self.sp_cm)):
            val = self.sp_cm[k]
            self.m1 = (((float(ws.cell(row=val, column=3).value) - float(cy)) ** 2) + (
                    (float(ws.cell(row=val, column=4).value) - float(cx)) ** 2) + ((float(ws.cell(row=val, column=5).value) - float(cm)) ** 2)) ** 0.5
            if self.m1 < self.m:
                self.m = self.m1
                self.k1 = val

        #вывод
        print("Наиближайший профиль")
        for n in range(1,6): #найденный номер наиближайшей строки хранится в k1, выводим значения из строки k1 и столбцов с коэффициентами
            print(ws.cell(row=self.k1, column=n).value)
        self.m = 9999.9 #возвращаем крупное значение для новых прогонов
        self.k1 = 1 #аналогично, чтобы при следующих прогонах не использовались старые данные
        self.m1 = 0

example = Profiles(cy=float, cx=float, cm=float)

example.relating(cy=float(input()), cx=float(input()), cm=float(input()))
example.sravn(cy=float(input()), cx=float(input()), cm=float(input()), val=int)
#print(example.relating(cy=float(input()), cx=float(input()), cm=float(input())))
