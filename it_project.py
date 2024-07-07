import matplotlib
import csv
import openpyxl
import pandas as pd

#вводим необходимые коэффициенты
cy = float(input())
cx = float(input())
cm = float(input())

#открываем в питоне эксель с базой данных и лист
wb = openpyxl.load_workbook('data.xlsx')
ws = wb.active


#пробегаем с 2 по 20 строчку, в каждой строчке пробегаем все столбцы
for i in range(2,20):
    for j in range(1,6):
        if cy == float(ws.cell(row=i, column=3).value) and cx == float(ws.cell(row=i, column=4).value) and cm == float(ws.cell(row=i, column=5).value):
            print(ws.cell(row=1,column=j).value, ws.cell(row=i, column=j).value) #если есть полное совпадение коэффициентов, выводим 1 столбец(название профиля) и все его коэффициенты
        else:
            break #если не подходит, завершаем искать в этой строчке и переходим на новую
